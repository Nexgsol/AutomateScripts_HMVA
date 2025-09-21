from typing import Optional
import uuid
from django.shortcuts import render, get_object_or_404, redirect
from core.jobs import job_get_or_create, job_set_state
from core.utils import generate_heritage_paragraph, call_openai_for_ssml, build_prompt,parse_openai_json
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import Brand, ScriptRequest, Template, AvatarProfile
from .serializers import ScriptRequestSerializer

from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from .adapters import avatar_heygen

from .tasks import (
    task_generate_script, task_kickoff_chain, task_render_avatar, task_assemble_template,
    task_push_drive, task_generate_captions, task_sync_airtable,
    task_schedule, task_publish, task_metrics_24h, orchestrate_paragraphs_job
)
from rest_framework.parsers import JSONParser, MultiPartParser, FormParser

def script_form(request):
    brands = Brand.objects.all()
    avatars = AvatarProfile.objects.all()
    templates = Template.objects.all()
    ctx = {"brands": brands, "avatars": avatars, "templates": templates}
    if request.method == "POST":
        brand_id = request.POST["brand"]
        icon = request.POST["icon"]
        notes = request.POST.get("notes","")
        duration = request.POST.get("duration","30s")
        avatar_id = request.POST.get("avatar") or None
        template_id = request.POST.get("template") or None
        sr = ScriptRequest.objects.create(
            brand_id=brand_id, icon_or_topic=icon, notes=notes,
            duration=duration, avatar_id=avatar_id, template_id=template_id, status="New"
        )
        if request.POST.get("auto") == "on":
            task_kickoff_chain.delay(sr.id)
            return redirect("request-detail", pk=sr.id)
        task_generate_script.delay(sr.id)
        return redirect("request-detail", pk=sr.id)
    return render(request, "script_form.html", ctx)


class AutoPipelineAPI(APIView):
    def post(self, req, pk):
        task_kickoff_chain.delay(pk)
        return Response({"id": pk, "auto_pipeline": "queued"}, status=status.HTTP_202_ACCEPTED)

def request_detail(request, pk):
    sr = get_object_or_404(ScriptRequest, pk=pk)
    return render(request, "request_detail.html", {"sr": sr})


# class ScriptGenerateAPI(APIView):
#     def post(self, req):
#         ser = ScriptRequestSerializer(data=req.data)
#         ser.is_valid(raise_exception=True)
#         sr = ser.save(status="New")
#         task_generate_script.delay(sr.id)
#         return Response({"id": sr.id, "status": "queued"}, status=status.HTTP_201_CREATED)

# class ScriptGetAPI(APIView):
#     def get(self, req, pk):
#         sr = get_object_or_404(ScriptRequest, pk=pk)
#         return Response(ScriptRequestSerializer(sr).data)

# class RenderAvatarAPI(APIView):
#     def post(self, req, pk):
#         sr = get_object_or_404(ScriptRequest, pk=pk)
#         task_render_avatar.delay(sr.id)
#         return Response({"id": sr.id, "render": "queued"}, status=status.HTTP_202_ACCEPTED)

# class AssembleAPI(APIView):
#     def post(self, req, pk):
#         sr = get_object_or_404(ScriptRequest, pk=pk)
#         task_assemble_template.delay(sr.id)
#         return Response({"id": sr.id, "assemble": "queued"}, status=status.HTTP_202_ACCEPTED)

# class DriveAPI(APIView):
#     def post(self, req, pk):
#         sr = get_object_or_404(ScriptRequest, pk=pk)
#         task_push_drive.delay(sr.id)
#         return Response({"id": sr.id, "drive": "queued"}, status=status.HTTP_202_ACCEPTED)

# class CaptionsAPI(APIView):
#     def post(self, req, pk):
#         sr = get_object_or_404(ScriptRequest, pk=pk)
#         task_generate_captions.delay(sr.id)
#         return Response({"id": sr.id, "captions": "queued"}, status=status.HTTP_202_ACCEPTED)

# class AirtableSyncAPI(APIView):
#     def post(self, req, pk):
#         sr = get_object_or_404(ScriptRequest, pk=pk)
#         task_sync_airtable.delay(sr.id)
#         return Response({"id": sr.id, "airtable": "queued"}, status=status.HTTP_202_ACCEPTED)

# class ScheduleAPI(APIView):
#     def post(self, req, pk):
#         sr = get_object_or_404(ScriptRequest, pk=pk)
#         task_schedule.delay(sr.id)
#         return Response({"id": sr.id, "schedule": "queued"}, status=status.HTTP_202_ACCEPTED)

# class PublishAPI(APIView):
#     def post(self, req, pk):
#         sr = get_object_or_404(ScriptRequest, pk=pk)
#         task_publish.delay(sr.id)
#         return Response({"id": sr.id, "publish": "queued"}, status=status.HTTP_202_ACCEPTED)

# class MetricsAPI(APIView):
#     def post(self, req, pk):
#         sr = get_object_or_404(ScriptRequest, pk=pk)
#         task_metrics_24h.delay(sr.id)
#         return Response({"id": sr.id, "metrics24h": "queued"}, status=status.HTTP_202_ACCEPTED)

# def avatar_quick_create(request):
#     brands = Brand.objects.all()
#     ctx = {"brands": brands}
#     if request.method == "POST":
#         brand_id = request.POST["brand"]
#         avatar_name = request.POST["avatar_name"]
#         voice_name = request.POST.get("voice_name","")
#         eleven_id = request.POST.get("eleven_id","")
#         img = request.FILES.get("photo")

#         image_path = None
#         if img:
#             image_path = default_storage.save(f"avatars/{img.name}", ContentFile(img.read()))
#             # Try optional HeyGen photo-avatar API (stub returns "")
#             avatar_id = avatar_heygen.create_photo_avatar(open(default_storage.path(image_path),"rb").read(), avatar_name)
#         else:
#             avatar_id = ""

#         ap = AvatarProfile.objects.create(
#             brand_id=brand_id,
#             avatar_name=avatar_name,
#             voice_name=voice_name,
#             elevenlabs_voice_id=eleven_id,
#             heygen_avatar_id=avatar_id,
#             image=image_path
#         )
#         return redirect("script-form")
#     return render(request, "avatar_form.html", ctx)


# class ParagraphAPI(APIView):
#     def post(self, request):
#         icon = request.data.get("icon") or request.POST.get("icon")
#         notes = request.data.get("notes") or request.POST.get("notes", "")
#         if not icon:
#             return Response({"error": "icon is required"}, status=status.HTTP_400_BAD_REQUEST)
#         paragraph = generate_heritage_paragraph(
#             icon, notes
#         )
#         return Response({"icon": icon, "paragraph": paragraph}, status=status.HTTP_200_OK)
    




from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.views.decorators.http import require_http_methods
from django.http import JsonResponse
from .forms import ScriptAvatarForm
from .models import Icon, ScriptRequest
from .utils import generate_heritage_paragraph
from .adapters import avatar_heygen
from .tasks import task_render_heygen_tts

@require_http_methods(["GET", "POST"])
def script_avatar_page(request):
    paragraph = ""
    if request.method == "POST":
        form = ScriptAvatarForm(request.POST)
        action = request.POST.get("action", "")
        if form.is_valid():
            icon_obj = form.cleaned_data["icon"]
            brand = form.cleaned_data["brand"]
            duration = form.cleaned_data["duration"]
            category = form.cleaned_data["category"] or (icon_obj.category if hasattr(icon_obj, "category") else "")
            notes = form.cleaned_data["notes"] or (getattr(icon_obj, "short_cues", "") or "")

            # Always refresh paragraph so both actions work
            paragraph = generate_heritage_paragraph(icon_obj.name, notes)

            if action == "generate_video":
                heygen_avatar_id = form.cleaned_data["heygen_avatar_id"]
                heygen_voice_id  = form.cleaned_data["heygen_voice_id"] or None
                if not heygen_avatar_id:
                    messages.error(request, "Please select a HeyGen avatar.")
                else:
                    sr = ScriptRequest.objects.create(
                        brand=brand, mode="Single",
                        icon_or_topic=icon_obj.name,
                        notes=notes,
                        duration=duration,
                        draft_script=paragraph,
                        final_script=paragraph,
                        status="Drafted",
                    )
                    task_render_heygen_tts.delay(sr.id, heygen_avatar_id, heygen_voice_id)
                    messages.success(request, f"Video render queued for {icon_obj.name}.")
                    # change 'request-detail' to your actual detail route name if different
                    return redirect("request-detail", pk=sr.id)
    else:
        form = ScriptAvatarForm()

    return render(request, "script_avatar_page.html", {"form": form, "paragraph": paragraph})


# AJAX helpers
def heygen_avatars_api(request):
    return JsonResponse({"avatars": avatar_heygen.list_avatars()})


def heygen_voices_api(request):
    return JsonResponse({"voices": avatar_heygen.list_voices()})


def icon_meta_api(request, pk: int):
    icon = get_object_or_404(Icon, pk=pk)
    return JsonResponse({"category": getattr(icon, "category", ""), "notes": getattr(icon, "short_cues", "")})


# views.py
import os
import json
import uuid
import requests
from django.conf import settings
from django.http import JsonResponse, HttpResponseBadRequest
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt  # not needed if you pass CSRF
from django.utils.text import get_valid_filename


@require_POST
def api_tts_elevenlabs(request):
    """
    POST: voice_id, ssml
    Returns: {"audio_url": "<MEDIA_URL>/tts/<file>.mp3"}
    """
    # Support both FormData and JSON
    voice_id = os.getenv("ELEVENLABS_VOICE_ID") or request.POST.get("voice_id")  # default from .env
    ssml = request.POST.get("ssml")
    if not voice_id or not ssml:
        try:
            data = json.loads(request.body or "{}")
            voice_id = os.getenv("ELEVENLABS_VOICE_ID") or data.get("voice_id")  # default from .env
            ssml = ssml or data.get("ssml")
        except Exception:
            pass

    if not voice_id or not ssml:
        return HttpResponseBadRequest("Missing voice_id or ssml")

    api_key = getattr(settings, "ELEVENLABS_API_KEY", None)
    if not api_key:
        return JsonResponse({"error": "ELEVENLABS_API_KEY not configured"}, status=500)

    # ElevenLabs TTS (stream) endpoint
    # Note: If your account requires "text" instead of "ssml",
    # you can switch to {"text": ssml, "use_sid": true} or their SSML flag.
    url = f"https://api.elevenlabs.io/v1/text-to-speech/{voice_id}/stream?output_format=mp3_44100_128"

    payload = {
        "model_id": "eleven_multilingual_v2",
        # IMPORTANT: SSML goes in "text"
        "text": ssml,
        # Tell ElevenLabs this is SSML
        "input_format": "ssml",
        "voice_settings": {
            "stability": 0.5,
            "similarity_boost": 0.75
        },
        # Optional but helpful
        "apply_text_normalization": "auto",
        "use_speaker_boost": True
    }

    headers = {
        "xi-api-key": api_key,
        "Accept": "audio/mpeg",
        "Content-Type": "application/json",
    }



    try:
        r = requests.post(url, headers=headers, json=payload, stream=True, timeout=120)
        if r.status_code != 200:
            # Some orgs get 422 if SSML flag/field differs — include server message
            return JsonResponse({"error": f"ElevenLabs error {r.status_code}: {r.text}"}, status=400)

        # Save MP3 under MEDIA_ROOT/tts/
        outdir = os.path.join(settings.MEDIA_ROOT, "tts")
        os.makedirs(outdir, exist_ok=True)
        filename = f"{uuid.uuid4().hex}.mp3"
        filepath = os.path.join(outdir, get_valid_filename(filename))

        with open(filepath, "wb") as f:
            for chunk in r.iter_content(chunk_size=8192):
                if chunk:
                    f.write(chunk)

        audio_url = f"{settings.MEDIA_URL}tts/{filename}"
        return JsonResponse({"audio_url": audio_url})
    except requests.RequestException as e:
        return JsonResponse({"error": f"Network error: {e}"}, status=502)


# Paragraph generator API (used by the left-side button)


# class ParagraphAPI(APIView):
#     """
#     POST multipart/form-data:
#       - file: .xlsx upload (required)
#       - sheet: optional (name or index)
#       - batch_size: optional (default 25)

#     Behavior:
#       - Saves the file
#       - Queues Celery job to process in batches
#       - Returns { job_id } with 202
#     """
#     parser_classes = (MultiPartParser, JSONParser, FormParser)

#     def post(self, request):
#         try:
#             if "file" not in request.FILES:
#                 return Response({"error": "No file provided"}, status=status.HTTP_400_BAD_REQUEST)

#             uploaded = request.FILES["file"]
#             sheet = request.data.get("sheet")
#             try:
#                 batch_size = int(request.data.get("batch_size", 25))
#                 if batch_size <= 0:
#                     raise ValueError
#             except Exception:
#                 return Response({"error": "batch_size must be a positive integer"}, status=400)

#             # Persist the upload so workers can read it
#             saved_path = default_storage.save(
#                 f"uploads/{uuid.uuid4()}_{uploaded.name}",
#                 uploaded,
#             )

#             # Enqueue orchestration task (to be implemented next)

#             task = orchestrate_paragraphs_job.delay(

#                 file_path=saved_path,
#                 sheet=sheet,
#                 batch_size=batch_size,
#             )

#             return Response({"job_id": task.id, "status": "queued"}, status=status.HTTP_202_ACCEPTED)

#         except Exception as e:
#             return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



from uuid import uuid4
import os
from django.core.files.storage import default_storage
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.parsers import MultiPartParser, JSONParser, FormParser

# your generator
from core.utils import generate_heritage_paragraph_with_ssml
# your Celery orchestration task (must accept the kwargs used below)



def _is_xlsx_upload(name: str, content_type: Optional[str]) -> bool:
    if not name:
        return False
    name_ok = name.lower().endswith(".xlsx")
    ct_ok = (content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    # some browsers miss content_type; file extension check is primary
    return name_ok or ct_ok


class ParagraphAPI(APIView):
    """
    Modes (by 'action'):

    A) action=upload_sheet  (multipart/form-data)
       - file: .xlsx (required)
       - sheet: optional (default 'Sheet1')
       - batch_size: optional int (default 25)
       -> Enqueues Celery orchestrator in mode='local_file'
       <- 202 { job_id, status:'queued', mode, file, sheet, batch_size, status_url }

    B) action=process_google_sheet  (JSON or form)
       - sheet_public_url (required, CSV export URL)
       - sheet_id (required)
       - sheet_name (default 'Sheet1')
       - batch_size: optional int (default 25)
       -> Enqueues mode='google_sheet'
       <- 202 { job_id, status:'queued', ... , status_url }

    C) default: single generate now
       - icon (required), notes/category/duration (optional)
       <- 200 { icon, data: { paragraph, ssml, ... } }
    """
    parser_classes = (MultiPartParser, JSONParser, FormParser)

    def post(self, request):
        action = (request.data.get("action") or request.POST.get("action") or "").strip()

        # ---------------- A) local .xlsx upload ----------------
        if action == "upload_sheet":
            if "file" not in request.FILES:
                return Response({"error": "No file provided"}, status=status.HTTP_400_BAD_REQUEST)

            uploaded = request.FILES["file"]
            if not _is_xlsx_upload(uploaded.name, uploaded.content_type):
                return Response({"error": "Please upload a valid .xlsx file"}, status=status.HTTP_400_BAD_REQUEST)

            sheet = request.data.get("sheet") or "Sheet1"
            try:
                batch_size = int(request.data.get("batch_size", 25))
                if batch_size <= 0:
                    raise ValueError
            except Exception:
                return Response({"error": "batch_size must be a positive integer"}, status=status.HTTP_400_BAD_REQUEST)

            # Save file where workers can read it
            job_id = str(uuid.uuid4())
            saved_path = default_storage.save(f"uploads/{job_id}_{uploaded.name}", uploaded)

            # PRE-CREATE JobRun so /api/jobs/ shows it immediately
            job_get_or_create(
                job_id,
                mode="local_file",
                file_path=saved_path,
                sheet_name=sheet,
            )
            job_set_state(job_id, state="PENDING")

            # Enqueue orchestrator (bind this API's job_id to the Celery id)
            task = orchestrate_paragraphs_job.apply_async(
                kwargs={
                    "mode": "local_file",
                    "file_path": saved_path,
                    "sheet": sheet,           # IMPORTANT: pass 'sheet'
                    "batch_size": batch_size,
                },
                task_id=job_id,
            )

            return Response(
                {
                    "job_id": task.id,
                    "status": "queued",
                    "mode": "local_file",
                    "file": saved_path,
                    "sheet": sheet,
                    "batch_size": batch_size,
                    "status_url": f"/api/jobs/{task.id}/status/",
                },
                status=status.HTTP_202_ACCEPTED,
            )

        # ---------------- B) Google Sheet (CSV URL) -------------
        if action == "process_google_sheet":
            sheet_public_url = request.data.get("sheet_public_url")
            gs_id = request.data.get("sheet_id")
            sheet_name = request.data.get("sheet_name") or "Sheet1"
            try:
                batch_size = int(request.data.get("batch_size", 25))
                if batch_size <= 0:
                    raise ValueError
            except Exception:
                return Response({"error": "batch_size must be a positive integer"}, status=status.HTTP_400_BAD_REQUEST)
            if not sheet_public_url or not gs_id:
                return Response({"error": "sheet_public_url and sheet_id are required"}, status=status.HTTP_400_BAD_REQUEST)

            # Create a friendly id and pre-create JobRun
            job_id = str(uuid.uuid4())
            job_get_or_create(
                job_id,
                mode="google_sheet",
                file_path="",
                sheet_name=sheet_name,
            )
            job_set_state(job_id, state="PENDING")

            # Enqueue with our chosen job_id
            task = orchestrate_paragraphs_job.apply_async(
                kwargs={
                    "mode": "google_sheet",
                    "sheet_public_url": sheet_public_url,
                    "sheet_id": gs_id,
                    "sheet_name": sheet_name,
                    "batch_size": batch_size,
                },
                task_id=job_id,
            )

            return Response(
                {
                    "job_id": task.id,
                    "status": "queued",
                    "mode": "google_sheet",
                    "sheet_id": gs_id,
                    "sheet_name": sheet_name,
                    "batch_size": batch_size,
                    "status_url": f"/api/jobs/{task.id}/status/",
                },
                status=status.HTTP_202_ACCEPTED,
            )

        # ---------------- C) single generate now ----------------
        icon = request.data.get("icon") or request.POST.get("icon")
        notes = request.data.get("notes") or request.POST.get("notes", "")
        category = request.data.get("category") or request.POST.get("category", "")
        duration = request.data.get("duration") or request.POST.get("duration", "")

        if not icon:
            return Response({"error": "icon is required"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            data = generate_heritage_paragraph_with_ssml(icon, notes, category)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response({"icon": icon, "data": data}, status=status.HTTP_200_OK)

# views.py
from django.http import JsonResponse, Http404
from celery.result import AsyncResult
from django.core.files.storage import default_storage
from openpyxl import load_workbook
from core.models import JobRun

RESULTS_DIR = "uploads/results"

TERMINAL = {"SUCCESS", "FAILURE"}

def job_status(request, job_id):
    """
    DB-first status:
    1) If JobRun exists and is terminal -> return DB state & payload.
    2) Else if handoff_id exists -> ask Celery final task; fall back to DB if Celery says PENDING/unknown.
    3) Else -> ask Celery orchestrator; if it returns a handoff OR DB becomes terminal, surface that.
    """
    job_id = str(job_id)
    out = {"job_id": job_id}

    jr = JobRun.objects.filter(job_id=job_id).first()

    # 1) Prefer terminal DB state (handles cases where Celery backend is not visible to web)
    if jr and jr.state in TERMINAL:
        out["state"] = jr.state
        if jr.state == "SUCCESS":
            out["download_url"] = jr.download_url or None
            out["results_file"] = jr.results_path or None
            out["mode"] = jr.mode or None
        else:
            out["error"] = jr.error or "Job failed."
        return JsonResponse(out)

    # 2) If we have a final callback id, try that first
    if jr and jr.handoff_id:
        final = AsyncResult(jr.handoff_id)
        state = getattr(final, "state", "PENDING")
        # If Celery backend is slow/unavailable, fall back to DB state
        if state in TERMINAL:
            out["state"] = state
            if state == "SUCCESS":
                fin = (final.result or {}) if isinstance(final.result, dict) else {}
                out.update({
                    "download_url": fin.get("download_url") or jr.download_url or None,
                    "results_file": fin.get("results") or jr.results_path or None,
                    "mode": fin.get("mode") or jr.mode or None,
                })
            else:
                out["error"] = str(final.result)
            return JsonResponse(out)
        # Non-terminal from Celery; surface non-terminal but include whatever we know from DB
        out["state"] = state
        if jr.download_url:
            out["download_url"] = jr.download_url
        if jr.results_path:
            out["results_file"] = jr.results_path
        if jr.mode:
            out["mode"] = jr.mode
        return JsonResponse(out)

    # 3) Fallback: orchestrator AsyncResult
    orch = AsyncResult(job_id)
    state = getattr(orch, "state", "PENDING")
    out["state"] = state

    if state == "SUCCESS":
        data = orch.result or {}
        handoff_id = data.get("handoff_id")
        if handoff_id:
            final = AsyncResult(handoff_id)
            fstate = getattr(final, "state", "PENDING")
            out["state"] = fstate
            if fstate == "SUCCESS":
                fin = final.result or {}
                out.update({
                    "download_url": fin.get("download_url") or (jr.download_url if jr else None),
                    "results_file": fin.get("results") or (jr.results_path if jr else None),
                    "mode": fin.get("mode") or (jr.mode if jr else None),
                })
            elif fstate == "FAILURE":
                out["error"] = str(final.result)
        else:
            # legacy path where orchestrator returns final payload
            out.update({
                "download_url": data.get("download_url") or (jr.download_url if jr else None),
                "results_file": data.get("results") or (jr.results_path if jr else None),
                "mode": data.get("mode") or (jr.mode if jr else None),
            })
        return JsonResponse(out)

    if state == "FAILURE":
        out["error"] = str(orch.result)
        return JsonResponse(out)

    # Non-terminal orchestrator; enrich with any DB info we have
    if jr:
        if jr.download_url:
            out["download_url"] = jr.download_url
        if jr.results_path:
            out["results_file"] = jr.results_path
        if jr.mode:
            out["mode"] = jr.mode
    return JsonResponse(out)



def _results_rel(job_id: str) -> str:
    return f"{RESULTS_DIR}/{job_id}.xlsx"


def job_results(request, job_id):
    job_id = str(job_id)
    results_rel = _results_rel(job_id)
    if not default_storage.exists(results_rel):
        raise Http404("Results file not found yet.")
    abs_path = default_storage.path(results_rel)

    wb = load_workbook(abs_path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    try:
        headers = list(next(rows_iter)) or []
    except StopIteration:
        headers = []

    rows = []
    limit = int(request.GET.get("limit", 200))
    for i, tup in enumerate(rows_iter, start=1):
        rows.append({str(headers[j] if j < len(headers) else f"col{j+1}"): tup[j] for j in range(len(tup))})
        if i >= limit:
            break
    wb.close()

    try:
        download_url = default_storage.url(results_rel)
    except Exception:
        download_url = None

    return JsonResponse({
        "job_id": job_id,
        "download_url": download_url,
        "headers": headers,
        "rows": rows,
        "count": len(rows),
    })
