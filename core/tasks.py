import os
import uuid
import json
import logging
import datetime

from celery import shared_task, group, chord, chain
from django.utils import timezone
from django.conf import settings
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage

from openpyxl import Workbook, load_workbook

from core.models import ScriptRequest, PublishTarget
from core.prompts import (
    GENERATOR_SYSTEM,
    gen_user,
    finalize_user,
    CAPTION_SYSTEM,
    captions_user,
)
from core.services.tts_service import fetch_or_create_tts_audio, load_audio_bytes
from core.adapters import (
    tts_elevenlabs,
    avatar_heygen,
    renderer_shotstack,
    renderer_cloudinary,
    airtable,
    drive_google,
    publish_youtube,
    publish_instagram,
    publish_facebook,
    publish_tiktok,
)
from core.utils import (
    build_prompt,
    parse_openai_json,
    call_openai_for_paragraph_and_ssml,
    iter_rows_streaming,
    batch,
)

logger = logging.getLogger(__name__)

RESULTS_DIR = "results"


@shared_task
def task_generate_script(sr_id: int):
    sr = ScriptRequest.objects.get(id=sr_id)
    lo, hi = utils.word_range(sr.duration)
    draft = utils.llm_chat(GENERATOR_SYSTEM, gen_user(sr.icon_or_topic, sr.notes, lo, hi), 0.5)
    qc = utils.qc_local(draft, lo, hi)
    sr.draft_script = draft
    sr.qc_json = qc
    if qc["fix_needed"]:
        final_text = utils.llm_chat("You are a precise editor.",
                                    finalize_user(sr.icon_or_topic, sr.notes, lo, hi, draft), 0.4)
        sr.final_script = final_text
        sr.qc_json = utils.qc_local(final_text, lo, hi)
        sr.status = "Drafted" if not sr.qc_json["fix_needed"] else "NeedsFix"
    else:
        sr.final_script = draft
        sr.status = "Drafted"
    sr.updated_at = timezone.now()
    sr.save()
    return sr.id


@shared_task
def task_render_avatar(sr_id: int):
    sr = ScriptRequest.objects.get(id=sr_id)
    if not sr.final_script:
        task_generate_script(sr.id)
        sr.refresh_from_db()

    if not (sr.avatar and sr.avatar.heygen_avatar_id and sr.avatar.elevenlabs_voice_id):
        raise ValueError("Avatar profile with HeyGen avatar id and ElevenLabs voice id is required.")

    # Fetch or create (re-use) the TTS asset and get its bytes
    tts_rec = fetch_or_create_tts_audio(
        voice_id=sr.avatar.elevenlabs_voice_id,
        text=sr.final_script,
        settings={"stability": 0.5, "similarity_boost": 0.75},
        attach_history=True,
    )
    mp3 = load_audio_bytes(tts_rec)

    # Continue with HeyGen upload/render (unchanged) ...
    res = avatar_heygen.generate_from_audio_bytes(
        sr.avatar.heygen_avatar_id, mp3, title=f"{sr.icon_or_topic} · req#{sr.id}"
    )
    if res.get("status") != "completed":
        sr.status = "Assembling"
        sr.qc_json = {**(sr.qc_json or {}), "heygen_status": res}
        sr.save()
        return res

    sr.asset_url = res.get("video_url","") or sr.asset_url
    sr.edit_url  = res.get("share_url","") or sr.edit_url
    sr.status = "Rendered"
    sr.save()
    return {"video_url": sr.asset_url, "share_url": sr.edit_url, "eleven_history_id": tts_rec.eleven_history_id}


@shared_task
def task_assemble_template(sr_id: int):
    sr = ScriptRequest.objects.get(id=sr_id)
    if not sr.asset_url:
        raise ValueError("No avatar asset_url to assemble")
    if sr.template and sr.template.engine == "cloudinary":
        res = renderer_cloudinary.assemble(sr.template.payload_json, sr.asset_url)
    else:
        res = renderer_shotstack.assemble(sr.template.payload_json if sr.template else {}, sr.asset_url)
    sr.asset_url = res.get("asset_url", sr.asset_url)
    sr.edit_url = res.get("edit_url", sr.edit_url)
    today = datetime.datetime.utcnow().strftime("%Y%m%d")
    content_type = {"15s":"15sStory","30s":"30sReel","60s":"60sReel"}.get(sr.duration,"30sReel")
    sr.file_name = f"{sr.icon_or_topic.replace(' ','_')}_{today}_9x16_{content_type}.mp4"
    sr.status = "Rendered"
    sr.updated_at = timezone.now()
    sr.save()
    return sr.file_name

@shared_task
def task_push_drive(sr_id: int):
    sr = ScriptRequest.objects.get(id=sr_id)
    if not sr.asset_url:
        raise ValueError("No asset_url to push to Drive")
    folder_id = settings.GDRIVE_FOLDER_ID or ""
    try:
        meta = drive_google.upload_from_url(sr.asset_url, sr.file_name or "Heritage_Reel.mp4", folder_id)
        sr.edit_url = meta.get("webViewLink","") or sr.edit_url
    except Exception as e:
        sr.qc_json = {**(sr.qc_json or {}), "drive_error": str(e)}
    sr.updated_at = timezone.now()
    sr.save()
    return True

@shared_task
def task_generate_captions(sr_id: int):
    sr = ScriptRequest.objects.get(id=sr_id)
    hashtags = (sr.brand.hashtags or "").strip()
    hashtags_csv = hashtags if hashtags else ""
    resp = utils.llm_chat(CAPTION_SYSTEM, captions_user(sr.icon_or_topic, hashtags_csv), 0.2)
    import json
    try:
        data = json.loads(resp)
    except Exception:
        data = {"caption_yt": resp[:180], "caption_tt": resp[:180],
                "caption_ig_reels": resp[:180], "caption_ig_stories": resp[:180], "caption_fb_reels": resp[:180]}
    sr.caption_yt = data.get("caption_yt","")
    sr.caption_tt = data.get("caption_tt","")
    sr.caption_ig_reels = data.get("caption_ig_reels","")
    sr.caption_ig_stories = data.get("caption_ig_stories","")
    sr.caption_fb_reels = data.get("caption_fb_reels","")
    sr.updated_at = timezone.now()
    sr.save()
    return True

@shared_task
def task_sync_airtable(sr_id: int):
    sr = ScriptRequest.objects.get(id=sr_id)
    fields = {
        "RequestID": sr.id, "Brand": sr.brand.name, "IconOrTopic": sr.icon_or_topic,
        "Duration": sr.duration, "Status": sr.status, "AssetURL": sr.asset_url,
        "FileName": sr.file_name, "CaptionYT": sr.caption_yt, "CaptionTikTok": sr.caption_tt,
        "CaptionIGReels": sr.caption_ig_reels, "CaptionIGStories": sr.caption_ig_stories,
        "CaptionFBReels": sr.caption_fb_reels,
    }
    rid = airtable.push_record(fields)
    return rid

@shared_task
def task_schedule(sr_id: int):
    sr = ScriptRequest.objects.get(id=sr_id)
    slot_label, slot_dt = utils.next_post_slot(sr.brand.timezone, sr.brand.post_windows)
    sr.scheduled_slot = slot_label
    sr.publish_at = slot_dt.astimezone(timezone.utc)
    sr.status = "Scheduled"
    sr.updated_at = timezone.now()
    sr.save()
    return sr.publish_at.isoformat()

@shared_task
def task_publish(sr_id: int):
    sr = ScriptRequest.objects.get(id=sr_id)
    posts = {}
    for t in PublishTarget.objects.filter(brand=sr.brand, enabled=True):
        if t.platform == "yt":
            posts["yt"] = publish_youtube.schedule_upload(sr.asset_url, sr.icon_or_topic, sr.publish_at.isoformat())
        elif t.platform == "ig":
            posts["ig"] = publish_instagram.schedule_upload(sr.asset_url, sr.caption_ig_reels, sr.publish_at.isoformat())
        elif t.platform == "fb":
            posts["fb"] = publish_facebook.schedule_upload(sr.asset_url, sr.caption_fb_reels, sr.publish_at.isoformat())
        elif t.platform == "tt":
            posts["tt"] = publish_tiktok.schedule_upload(sr.asset_url, sr.caption_tt, sr.publish_at.isoformat())
    sr.post_ids = posts
    sr.status = "Posted"
    sr.updated_at = timezone.now()
    sr.save()
    return posts

@shared_task
def task_metrics_24h(sr_id: int):
    sr = ScriptRequest.objects.get(id=sr_id)
    perf = {"views": 0, "likes": 0, "comments": 0, "shares": 0, "rank_percentile": 50}
    sr.performance_json = perf
    sr.status = "Pulled"
    sr.updated_at = timezone.now()
    sr.save()
    return perf

from requests.exceptions import HTTPError
from django.utils import timezone
from celery import shared_task
from core.models import ScriptRequest
from core.adapters import avatar_heygen, tts_elevenlabs


def _resolve_character_and_voice(chosen_id: str, provided_voice_id: str | None):
    """
    Return (character_type, look_id, voice_id).
    - character_type: "talking_photo" for motion looks, "avatar" for photo looks
    - look_id: the exact id to send to HeyGen (avatar_id or talking_photo_id)
    - voice_id: provided one or the look's default
    We try to match the chosen_id against known looks; if it is a group id, we
    pick a look (prefer motion if available).
    """
    # 1) try to match chosen_id as a known look
    all_looks_list = avatar_heygen.list_avatars()
    all_looks = {a.get("id") or a.get("avatar_id"): a for a in all_looks_list}
    look = all_looks.get(chosen_id)

    if look:
        is_motion = bool(look.get("is_motion"))
        ctype = "talking_photo" if is_motion else "avatar"
        v_id = provided_voice_id or look.get("default_voice_id") or None
        # prefer explicit avatar_id if present
        look_id = look.get("avatar_id") or look.get("id") or chosen_id
        return ctype, look_id, v_id

@shared_task
def task_render_heygen_tts(sr_id: int, avatar_or_group_id: str, heygen_voice_id: str | None = None):
    """
    Render via HeyGen using TTS first; if HeyGen rejects, fall back to audio pipeline.
    Automatically chooses avatar vs talking_photo based on the selected look.
    """
    sr = ScriptRequest.objects.get(id=sr_id)

    # Ensure we have text
    if not sr.final_script:
        sr.final_script = avatar_heygen.generate_heritage_paragraph(sr.icon_or_topic, sr.notes or "")
        sr.status = "Drafted"
        sr.save()

    # Decide character type + final look id + voice
    character_type, look_id, picked_voice = _resolve_character_and_voice(avatar_or_group_id, heygen_voice_id)
    print("type", character_type)

    # --- Try TTS path first ---
    try:
        if character_type == "avatar":
            video_id = avatar_heygen.create_avatar_video_from_text(
                avatar_id=look_id,
                input_text=sr.final_script,
                voice_id=picked_voice,
                title=f"{sr.icon_or_topic} · req#{sr.id}",
                width=1080, height=1920,
                accept_group_id=False,   # already resolved to a look id
            )
        else:
            video_id = avatar_heygen.create_talking_photo_video_from_text(
                talking_photo_id=look_id,
                input_text=sr.final_script,
                voice_id=picked_voice,
                title=f"{sr.icon_or_topic} · req#{sr.id}",
                width=1080, height=1920,
            )
    except HTTPError as e:
        # --- Fallback: ElevenLabs TTS -> upload audio -> render via audio asset ---
        el_voice = getattr(sr.avatar, "elevenlabs_voice_id", None)
        mp3_bytes = tts_elevenlabs.synthesize_tts_bytes(sr.final_script, el_voice)  # see adapter change below
        audio_asset_id = avatar_heygen.upload_audio_asset(mp3_bytes)

        if character_type == "avatar":
            video_id = avatar_heygen.create_avatar_video_from_audio(
                avatar_id=look_id,
                audio_asset_id=audio_asset_id,
                title=f"{sr.icon_or_topic} · req#{sr.id}",
                width=1080, height=1920,
                accept_group_id=False,
            )
        else:
            video_id = avatar_heygen.create_talking_photo_video_from_audio(
                talking_photo_id=look_id,
                audio_asset_id=audio_asset_id,
                title=f"{sr.icon_or_topic} · req#{sr.id}",
                width=1080, height=1920,
            )

    # Wait and persist
    st = avatar_heygen.wait_for_video(video_id, timeout_sec=900, poll_sec=10)
    if st.get("status") != "completed":
        sr.status = "Assembling"
        sr.qc_json = {**(sr.qc_json or {}), "heygen_status": st}
        sr.save()
        return st

    sr.asset_url = st.get("video_url", "")
    sr.edit_url = avatar_heygen.get_share_url(video_id) or ""
    sr.status = "Rendered"
    sr.updated_at = timezone.now()
    sr.save()
    return {"video_url": sr.asset_url, "share_url": sr.edit_url}


@shared_task
def task_kickoff_chain(sr_id: int):
    flow = chain(
        # task_generate_script.si(sr_id),
        task_render_avatar.si(sr_id),
        task_assemble_template.si(sr_id),
        task_push_drive.si(sr_id),
        task_generate_captions.si(sr_id),
        task_sync_airtable.si(sr_id),
        task_schedule.si(sr_id),
    )
    flow.delay()
    return {"pipeline": "queued", "sr_id": sr_id}


@shared_task(bind=True, autoretry_for=(Exception,), retry_backoff=True, retry_kwargs={"max_retries": 3})
def process_row_task(self, row_dict: dict) -> dict:
    """
    row_dict: {"row": int, "icon": str, "category": str, "notes": str}
    Returns:  {"row", "icon", "category", "notes", "paragraph", "ssml"}
    """
    icon = row_dict.get("icon", "").strip()
    category = row_dict.get("category", "").strip()
    notes = row_dict.get("notes", "").strip()

    prompt = build_prompt(icon=icon, notes=notes, category=category)
    raw = call_openai_for_paragraph_and_ssml(prompt)
    paragraph, ssml = parse_openai_json(raw)

    return {
        "row": row_dict["row"],
        "icon": icon,
        "category": category,
        "notes": notes,
        "paragraph": paragraph,
        "ssml": ssml,
    }

@shared_task(bind=True)
def save_batch_task(self, batch_results: list, job_id: str, batch_no: int, results_path: str) -> dict:
    """
    Appends batch_results to an XLSX file instead of JSONL.
    Columns: row, icon, category, notes, paragraph, ssml
    """
    # Ensure deterministic order inside a batch
    batch_results = sorted(batch_results, key=lambda x: x.get("row", 0))

    abs_results = default_storage.path(results_path)

    if not os.path.exists(abs_results):
        # Create new workbook with headers
        wb = Workbook()
        ws = wb.active
        ws.append(["row", "icon", "category", "notes", "paragraph", "ssml"])
    else:
        wb = load_workbook(abs_results)
        ws = wb.active

    for item in batch_results:
        ws.append([
            item.get("row", ""),
            item.get("icon", ""),
            item.get("category", ""),
            item.get("notes", ""),
            item.get("paragraph", ""),
            item.get("ssml", ""),
        ])

    wb.save(abs_results)

    logger.info(f"[Batch {batch_no}] Saved {len(batch_results)} rows to {results_path}")

    return {"saved_batch": batch_no, "count": len(batch_results)}

@shared_task(bind=True)
def orchestrate_paragraphs_job(self, file_path: str, sheet=None, batch_size: int = 25) -> dict:
    job_id = self.request.id or str(uuid.uuid4())
    abs_path = default_storage.path(file_path)

    # Prepare XLSX file (with headers) instead of JSONL
    results_rel = f"{RESULTS_DIR}/{job_id}.xlsx"
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["row", "icon", "category", "notes", "paragraph", "ssml"])
    wb.save(default_storage.path(results_rel))

    total_batches = 0
    for total_batches, rows in enumerate(batch(iter_rows_streaming(abs_path, sheet=sheet), batch_size), start=1):
        g = group(process_row_task.s(row) for row in rows)
        chord(g)(save_batch_task.s(job_id, total_batches, results_rel))

        logger.info(f"Scheduled batch {total_batches} for job {job_id}")
        self.update_state(state="PROGRESS", meta={"scheduled_batches": total_batches})

    return {"job_id": job_id, "results": results_rel, "batches": total_batches}
