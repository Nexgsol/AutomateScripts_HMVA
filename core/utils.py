import json
import re
import datetime
from zoneinfo import ZoneInfo

from core.adapters import llm_openai
from core.prompts import PROMPT_TEMPLATE



def word_range(duration: str):
    return {"15s": (60,75), "30s": (90,120), "60s": (150,180)}.get(duration, (90,120))

def llm_chat(system, user, temp=0.5):
    return llm_openai.chat(system, user, temp)

def call_openai_for_ssml(prompt):
    """
    Calls OpenAI's API via llm_openai.chat to convert text to SSML format.
    Returns the SSML string.
    """
    system = """You are an assistant that converts plain text into VALID, production-ready SSML for speech synthesis (ElevenLabs-compatible).Hard requirements:
    - Return ONE <speak> block ONLY. No code fences, no explanations, no XML declaration.
    - Use <prosody rate="medium"> for the wrapper unless specified.
    - Use <break> with milliseconds (120–500ms) to create natural pacing between beats.
    - Use <emphasis level="moderate"> to highlight 1–3 key phrases only.
    - Convert years to <say-as interpret-as="date" format="y">YYYY</say-as> and integers to <say-as interpret-as="cardinal">N</say-as> where appropriate.
    - Keep sentences 8–22 words for rhythm. Vary lengths slightly.
    - Do NOT invent content; preserve meaning and order. Lightly segment long sentences for clarity.
    - Escape special characters (&, <, >) if present in the input.
    - End with <mark name="END"/> just before closing </speak>.
    - No <audio> tags, no SSML comments, no vendor-specific tags.

    Voice guidance (implicit, do not output as text):
    - Tone: modern, confident, understated.
    - Diction: clean and warm; avoid hype.

    If the input already contains SSML, rebuild it into a single clean, standards-compliant block following the same rules.
    """
    user = prompt
    try:
        ssml = llm_openai.chat(system, user, temperature=0.2)
        return ssml
    except Exception as e:
        return f"Error: {str(e)}"

def count_words(t): return len(re.findall(r"\b[\w’']+\b", t))
def first_sentence(t):
    parts = re.split(r"(?<=[.!?])\s+", t.strip(), 1)
    return parts[0] if parts else t.strip()

def qc_local(text: str, lo: int, hi: int):
    wc = count_words(text)
    fs_wc = count_words(first_sentence(text))
    hook_ok = fs_wc <= 18
    punct_ok = ("—" not in text) and ("–" not in text) and not re.search(r"[😀-🙏]", text)
    risk = []
    if not punct_ok: risk.append("punctuation")
    if not hook_ok: risk.append("hook_long")
    if wc < lo or wc > hi: risk.append("length_out_of_range")
    has_six = hook_ok and (lo <= wc <= hi)
    return {"word_count": wc,"first_sentence_word_count": fs_wc,"hook_ok": hook_ok,
            "has_six_beats": has_six,"punctuation_rules_ok": punct_ok,
            "risk_flags": risk,"fix_needed": bool(risk) or not has_six,"suggested_edit": ""}

def next_post_slot(brand_timezone: str, post_windows_csv: str):
    tz = ZoneInfo(brand_timezone or "America/New_York")
    now = datetime.datetime.now(tz)
    windows = [w.strip() for w in post_windows_csv.split(",") if w.strip()]
    today_slots = []
    for w in windows:
        h, m = map(int, w.split(":"))
        slot = now.replace(hour=h, minute=m, second=0, microsecond=0)
        if slot > now:
            today_slots.append(slot)
    if today_slots:
        slot = min(today_slots)
    else:
        h, m = map(int, windows[0].split(":"))
        slot = (now + datetime.timedelta(days=1)).replace(hour=h, minute=m, second=0, microsecond=0)
    return slot.strftime("%H:%M"), slot


# core/utils.py (append at bottom or near your other llm helpers)
import re

def generate_heritage_paragraph(icon_name: str, notes: str) -> str:
    """
    Generates a single 100–130 word paragraph per the Instructional Notes.
    Uses llm_chat(BASE_SCRIPT_SYSTEM, base_script_user, temperature=0.5).
    Enforces 'single paragraph' and trims stray newlines/spaces.
    """
    from .prompts import BASE_SCRIPT_SYSTEM, base_script_user

    raw = llm_chat(BASE_SCRIPT_SYSTEM, base_script_user(icon_name, notes), temp=0.5)

    # normalize whitespace to keep one paragraph
    text = raw.strip()
    # Replace hard newlines with spaces; collapse multiple spaces
    text = re.sub(r'\s*\n+\s*', ' ', text)
    text = re.sub(r'\s{2,}', ' ', text).strip()

    # Guard: if it somehow produced multiple paragraphs, squash to one
    parts = [p.strip() for p in re.split(r'\n{2,}', text) if p.strip()]
    if len(parts) > 1:
        text = ' '.join(parts)

    # (Optional) hard length nudge: if outside 100–130 words, ask model to tighten/expand once.
    words = len(text.split())
    if words < 100 or words > 130:
        fix_prompt = (
            f"Rewrite to a single flowing paragraph of 100 to 130 words, keep meaning and all six beats, "
            f"no emojis, no em dashes, standard punctuation only. Icon: {icon_name}. Notes: {notes}. "
            f"ORIGINAL:\n{text}\n\nReturn only the corrected paragraph."
        )
        text = llm_chat("You are a precise editor.", fix_prompt, temp=0.3).strip()
        text = re.sub(r'\s*\n+\s*', ' ', text)
        text = re.sub(r'\s{2,}', ' ', text).strip()
    return text

def compose_icon_for_prompt(icon: str, category: str | None = None) -> str:
    icon = (icon or "").strip()
    category = (category or "").strip()
    return f"{icon} ({category})" if icon and category else icon

def build_prompt(icon: str, notes: str = "", category: str | None = None) -> str:
    icon_for_prompt = compose_icon_for_prompt(icon, category)
    return PROMPT_TEMPLATE.format(icon=icon_for_prompt, notes=(notes or "").strip())

def parse_openai_json(raw: str) -> tuple[str, str]:
    try:
        data = json.loads(raw)
        return str(data.get("paragraph", "")), str(data.get("ssml", ""))
    except Exception:
        return (raw or ""), ""

def iter_rows_streaming(file_like_or_path, sheet=None):
    """
    Stream normalized rows from a large .xlsx without loading the whole sheet.
    Yields dicts: {"row": excel_row_number, "icon", "category", "notes"}.
    - file_like_or_path: file path or Django UploadedFile / file-like object
    - sheet: None=first sheet, or sheet index, or sheet name
    """
    from openpyxl import load_workbook

    # open from file path or file-like
    if hasattr(file_like_or_path, "read"):
        fh = file_like_or_path
        try:
            fh.seek(0)
        except Exception:
            pass
        wb = load_workbook(fh, read_only=True, data_only=True)
    else:
        wb = load_workbook(file_like_or_path, read_only=True, data_only=True)

    ws = wb.worksheets[0] if sheet is None else (wb.worksheets[sheet] if isinstance(sheet, int) else wb[sheet])

    def _canon(s: str) -> str:
        return (s or "").strip().lower().replace("_", " ").replace("-", " ")

    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        wb.close()
        return

    hmap = { _canon(h): i for i, h in enumerate(header or []) if h is not None }

    def _get(row, *cands):
        """
        Return a cell value from `row` by checking candidate column names.

        Args:
            row (tuple): Row values from Excel.
            *cands (str): Possible header names (e.g. "Icon Name", "Icon").

        Returns:
            str: Normalized string from the first matching column,
                or "" if none found.
        """
        for c in cands:
            idx = hmap.get(_canon(c))
            if idx is not None:
                return str(row[idx] or "").strip()
        return ""


    excel_row = 2  # header is row 1
    for r in rows:
        icon = _get(r, "icon name", "icon", "name", "ICon name")
        if icon:
            yield {
                "row": excel_row,
                "icon": icon,
                "category": _get(r, "category", "type"),
                "notes": _get(r, "notes", "note"),
            }
        excel_row += 1

    wb.close()


def batch(iterable, size: int):
    """Yield lists of up to `size` items from an iterator (memory-light)."""
    buf = []
    for item in iterable:
        buf.append(item)
        if len(buf) == size:
            yield buf
            buf = []
    if buf:
        yield buf


def call_openai_for_paragraph_and_ssml(prompt: str) -> str:
    """
    Calls OpenAI once with both paragraph + SSML instructions.
    Returns raw JSON string (the model must output strictly JSON).
    """
    system = (
        "You are a senior fashion copywriter AND an SSML engineer.\n"
        "Return ONLY a single JSON object (no extra commentary, no markdown)."
    )
    try:
        raw = llm_openai.chat(system, prompt, temperature=0.2).strip()
        return raw
    except Exception as e:
        return json.dumps({
            "paragraph": "",
            "ssml": "",
            "error": str(e),
        })
