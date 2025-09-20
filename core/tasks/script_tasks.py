# core/tasks.py
from __future__ import annotations

import os
import io
import uuid
import json
import logging
import datetime
from typing import Dict, Iterable, List, Optional

import pandas as pd
import requests
from celery import shared_task, group, chord, chain
from django.conf import settings
from django.core.files.storage import default_storage
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from requests.exceptions import HTTPError

from core.models import ScriptRequest, PublishTarget
from core.prompts import (
    GENERATOR_SYSTEM,
    gen_user,
    finalize_user,
    CAPTION_SYSTEM,
    captions_user,
)
from core.services.tts_service import fetch_or_create_tts_audio, load_audio_bytes
from core.adapters import (
    tts_elevenlabs,
    avatar_heygen,
    renderer_shotstack,
    renderer_cloudinary,
    airtable,
    drive_google,
    publish_youtube,
    publish_instagram,
    publish_facebook,
    publish_tiktok,
)
from core import utils
from core.utils import (
    build_prompt,
    parse_openai_json,
    call_openai_for_paragraph_and_ssml,
    iter_rows_streaming,
)

# ---- Google Sheets helpers
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build

logger = logging.getLogger(__name__)

RESULTS_DIR = "uploads/results"  

# ====================== Small helpers ======================

def _a1_col(n: int) -> str:
    """1 -> A, 26 -> Z, 27 -> AA ..."""
    s = ""
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s

def _sheets_client():
    sa_path = os.environ.get("GOOGLE_SA_JSON") or getattr(settings, "GOOGLE_SA_JSON", None)
    if not sa_path:
        raise RuntimeError("GOOGLE_SA_JSON is not set for Google Sheets write-back.")
    creds = Credentials.from_service_account_file(sa_path, scopes=["https://www.googleapis.com/auth/spreadsheets"])
    return build("sheets", "v4", credentials=creds).spreadsheets()

def _ensure_sheet_headers_and_map(sheets, sheet_id: str, sheet_name: str, needed_cols: List[str]) -> Dict[str, int]:
    """
    Ensure 'needed_cols' appear in the first row. Return a 1-based name->index map.
    """
    head_resp = sheets.values().get(spreadsheetId=sheet_id, range=f"{sheet_name}!1:1").execute()
    header = head_resp.get("values", [[]])
    header = header[0] if header else []

    changed = False
    for col in needed_cols:
        if col not in header:
            header.append(col)
            changed = True
    if changed:
        sheets.values().update(
            spreadsheetId=sheet_id,
            range=f"{sheet_name}!A1:{_a1_col(len(header))}1",
            valueInputOption="RAW",
            body={"values": [header]},
        ).execute()

    return {name: i + 1 for i, name in enumerate(header)}

# ====================== Script & Render pipeline ======================

@shared_task
def task_generate_script(sr_id: int):
    sr = ScriptRequest.objects.get(id=sr_id)
    lo, hi = utils.word_range(sr.duration)
    draft = utils.llm_chat(GENERATOR_SYSTEM, gen_user(sr.icon_or_topic, sr.notes, lo, hi), 0.5)
    qc = utils.qc_local(draft, lo, hi)
    sr.draft_script = draft
    sr.qc_json = qc
    if qc["fix_needed"]:
        final_text = utils.llm_chat(
            "You are a precise editor.",
            finalize_user(sr.icon_or_topic, sr.notes, lo, hi, draft),
            0.4,
        )
        sr.final_script = final_text
        sr.qc_json = utils.qc_local(final_text, lo, hi)
        sr.status = "Drafted" if not sr.qc_json["fix_needed"] else "NeedsFix"
    else:
        sr.final_script = draft
        sr.status = "Drafted"
    sr.updated_at = timezone.now()
    sr.save()
    return sr.id


@shared_task
def task_render_avatar(sr_id: int):
    sr = ScriptRequest.objects.get(id=sr_id)
    if not sr.final_script:
        task_generate_script(sr.id)
        sr.refresh_from_db()

    if not (sr.avatar and sr.avatar.heygen_avatar_id and sr.avatar.elevenlabs_voice_id):
        raise ValueError("Avatar profile with HeyGen avatar id and ElevenLabs voice id is required.")

    tts_rec = fetch_or_create_tts_audio(
        voice_id=sr.avatar.elevenlabs_voice_id,
        text=sr.final_script,
        settings={"stability": 0.5, "similarity_boost": 0.75},
        attach_history=True,
    )
    mp3 = load_audio_bytes(tts_rec)

    res = avatar_heygen.generate_from_audio_bytes(
        sr.avatar.heygen_avatar_id, mp3, title=f"{sr.icon_or_topic} · req#{sr.id}"
    )
    if res.get("status") != "completed":
        sr.status = "Assembling"
        sr.qc_json = {**(sr.qc_json or {}), "heygen_status": res}
        sr.save()
        return res

    sr.asset_url = res.get("video_url", "") or sr.asset_url
    sr.edit_url = res.get("share_url", "") or sr.edit_url
    sr.status = "Rendered"
    sr.save()
    return {"video_url": sr.asset_url, "share_url": sr.edit_url, "eleven_history_id": tts_rec.eleven_history_id}


@shared_task
def task_assemble_template(sr_id: int):
    sr = ScriptRequest.objects.get(id=sr_id)
    if not sr.asset_url:
        raise ValueError("No avatar asset_url to assemble")

    if sr.template and sr.template.engine == "cloudinary":
        res = renderer_cloudinary.assemble(sr.template.payload_json, sr.asset_url)
    else:
        res = renderer_shotstack.assemble(sr.template.payload_json if sr.template else {}, sr.asset_url)

    sr.asset_url = res.get("asset_url", sr.asset_url)
    sr.edit_url = res.get("edit_url", sr.edit_url)

    today = datetime.datetime.utcnow().strftime("%Y%m%d")
    content_type = {"15s": "15sStory", "30s": "30sReel", "60s": "60sReel"}.get(sr.duration, "30sReel")
    sr.file_name = f"{sr.icon_or_topic.replace(' ', '_')}_{today}_9x16_{content_type}.mp4"
    sr.status = "Rendered"
    sr.updated_at = timezone.now()
    sr.save()
    return sr.file_name


@shared_task
def task_push_drive(sr_id: int):
    sr = ScriptRequest.objects.get(id=sr_id)
    if not sr.asset_url:
        raise ValueError("No asset_url to push to Drive")
    folder_id = settings.GDRIVE_FOLDER_ID or ""
    try:
        meta = drive_google.upload_from_url(sr.asset_url, sr.file_name or "Heritage_Reel.mp4", folder_id)
        sr.edit_url = meta.get("webViewLink", "") or sr.edit_url
    except Exception as e:
        sr.qc_json = {**(sr.qc_json or {}), "drive_error": str(e)}
    sr.updated_at = timezone.now()
    sr.save()
    return True


@shared_task
def task_generate_captions(sr_id: int):
    sr = ScriptRequest.objects.get(id=sr_id)
    hashtags = (sr.brand.hashtags or "").strip()
    hashtags_csv = hashtags if hashtags else ""
    resp = utils.llm_chat(CAPTION_SYSTEM, captions_user(sr.icon_or_topic, hashtags_csv), 0.2)
    try:
        data = json.loads(resp)
    except Exception:
        data = {
            "caption_yt": resp[:180],
            "caption_tt": resp[:180],
            "caption_ig_reels": resp[:180],
            "caption_ig_stories": resp[:180],
            "caption_fb_reels": resp[:180],
        }
    sr.caption_yt = data.get("caption_yt", "")
    sr.caption_tt = data.get("caption_tt", "")
    sr.caption_ig_reels = data.get("caption_ig_reels", "")
    sr.caption_ig_stories = data.get("caption_ig_stories", "")
    sr.caption_fb_reels = data.get("caption_fb_reels", "")
    sr.updated_at = timezone.now()
    sr.save()
    return True


@shared_task
def task_sync_airtable(sr_id: int):
    sr = ScriptRequest.objects.get(id=sr_id)
    fields = {
        "RequestID": sr.id,
        "Brand": sr.brand.name,
        "IconOrTopic": sr.icon_or_topic,
        "Duration": sr.duration,
        "Status": sr.status,
        "AssetURL": sr.asset_url,
        "FileName": sr.file_name,
        "CaptionYT": sr.caption_yt,
        "CaptionTikTok": sr.caption_tt,
        "CaptionIGReels": sr.caption_ig_reels,
        "CaptionIGStories": sr.caption_ig_stories,
        "CaptionFBReels": sr.caption_fb_reels,
    }
    return airtable.push_record(fields)


@shared_task
def task_schedule(sr_id: int):
    sr = ScriptRequest.objects.get(id=sr_id)
    slot_label, slot_dt = utils.next_post_slot(sr.brand.timezone, sr.brand.post_windows)
    sr.scheduled_slot = slot_label
    sr.publish_at = slot_dt.astimezone(timezone.utc)
    sr.status = "Scheduled"
    sr.updated_at = timezone.now()
    sr.save()
    return sr.publish_at.isoformat()


@shared_task
def task_publish(sr_id: int):
    sr = ScriptRequest.objects.get(id=sr_id)
    posts = {}
    for t in PublishTarget.objects.filter(brand=sr.brand, enabled=True):
        if t.platform == "yt":
            posts["yt"] = publish_youtube.schedule_upload(sr.asset_url, sr.icon_or_topic, sr.publish_at.isoformat())
        elif t.platform == "ig":
            posts["ig"] = publish_instagram.schedule_upload(
                sr.asset_url, sr.caption_ig_reels, sr.publish_at.isoformat()
            )
        elif t.platform == "fb":
            posts["fb"] = publish_facebook.schedule_upload(
                sr.asset_url, sr.caption_fb_reels, sr.publish_at.isoformat()
            )
        elif t.platform == "tt":
            posts["tt"] = publish_tiktok.schedule_upload(sr.asset_url, sr.caption_tt, sr.publish_at.isoformat())
    sr.post_ids = posts
    sr.status = "Posted"
    sr.updated_at = timezone.now()
    sr.save()
    return posts


@shared_task
def task_metrics_24h(sr_id: int):
    sr = ScriptRequest.objects.get(id=sr_id)
    perf = {"views": 0, "likes": 0, "comments": 0, "shares": 0, "rank_percentile": 50}
    sr.performance_json = perf
    sr.status = "Pulled"
    sr.updated_at = timezone.now()
    sr.save()
    return perf


def _resolve_character_and_voice(chosen_id: str, provided_voice_id: Optional[str]):
    all_looks_list = avatar_heygen.list_avatars()
    all_looks = {a.get("id") or a.get("avatar_id"): a for a in all_looks_list}
    look = all_looks.get(chosen_id)
    if look:
        is_motion = bool(look.get("is_motion"))
        ctype = "talking_photo" if is_motion else "avatar"
        v_id = provided_voice_id or look.get("default_voice_id") or None
        look_id = look.get("avatar_id") or look.get("id") or chosen_id
        return ctype, look_id, v_id
    return "avatar", chosen_id, provided_voice_id


@shared_task
def task_render_heygen_tts(sr_id: int, avatar_or_group_id: str, heygen_voice_id: Optional[str] = None):
    sr = ScriptRequest.objects.get(id=sr_id)

    if not sr.final_script:
        sr.final_script = avatar_heygen.generate_heritage_paragraph(sr.icon_or_topic, sr.notes or "")
        sr.status = "Drafted"
        sr.save()

    character_type, look_id, picked_voice = _resolve_character_and_voice(avatar_or_group_id, heygen_voice_id)

    try:
        if character_type == "avatar":
            video_id = avatar_heygen.create_avatar_video_from_text(
                avatar_id=look_id,
                input_text=sr.final_script,  # NOTE: plain text (not SSML)
                voice_id=picked_voice,       # HeyGen voice id (linked to 11L if needed)
                title=f"{sr.icon_or_topic} · req#{sr.id}",
                width=1080,
                height=1920,
                accept_group_id=False,
            )
        else:
            video_id = avatar_heygen.create_talking_photo_video_from_text(
                talking_photo_id=look_id,
                input_text=sr.final_script,
                voice_id=picked_voice,
                title=f"{sr.icon_or_topic} · req#{sr.id}",
                width=1080,
                height=1920,
            )
    except HTTPError:
        el_voice = getattr(sr.avatar, "elevenlabs_voice_id", None)
        mp3_bytes = tts_elevenlabs.synthesize_tts_bytes(sr.final_script, el_voice)
        audio_asset_id = avatar_heygen.upload_audio_asset(mp3_bytes)

        if character_type == "avatar":
            video_id = avatar_heygen.create_avatar_video_from_audio(
                avatar_id=look_id,
                audio_asset_id=audio_asset_id,
                title=f"{sr.icon_or_topic} · req#{sr.id}",
                width=1080,
                height=1920,
                accept_group_id=False,
            )
        else:
            video_id = avatar_heygen.create_talking_photo_video_from_audio(
                talking_photo_id=look_id,
                audio_asset_id=audio_asset_id,
                title=f"{sr.icon_or_topic} · req#{sr.id}",
                width=1080,
                height=1920,
            )

    st = avatar_heygen.wait_for_video(video_id, timeout_sec=900, poll_sec=10)
    if st.get("status") != "completed":
        sr.status = "Assembling"
        sr.qc_json = {**(sr.qc_json or {}), "heygen_status": st}
        sr.save()
        return st

    sr.asset_url = st.get("video_url", "")
    sr.edit_url = avatar_heygen.get_share_url(video_id) or ""
    sr.status = "Rendered"
    sr.updated_at = timezone.now()
    sr.save()
    return {"video_url": sr.asset_url, "share_url": sr.edit_url}


# @shared_task
# def task_render_heygen_tts(sr_id: int, avatar_or_group_id: str, heygen_voice_id: str | None = None):
#     """
#     Render via HeyGen using TTS first; if HeyGen rejects, fall back to audio pipeline.
#     Automatically chooses avatar vs talking_photo based on the selected look.
#     """
#     sr = ScriptRequest.objects.get(id=sr_id)

#     # Ensure we have text
#     if not sr.final_script:
#         sr.final_script = avatar_heygen.generate_heritage_paragraph(sr.icon_or_topic, sr.notes or "")
#         sr.status = "Drafted"
#         sr.save()

#     # Decide character type + final look id + voice
#     character_type, look_id, picked_voice = _resolve_character_and_voice(avatar_or_group_id, heygen_voice_id)
#     print("type", character_type)

#     # --- Try TTS path first ---
#     try:
#         if character_type == "avatar":
#             video_id = avatar_heygen.create_avatar_video_from_text(
#                 avatar_id=look_id,
#                 input_text=sr.final_script,
#                 voice_id=picked_voice,
#                 title=f"{sr.icon_or_topic} · req#{sr.id}",
#                 width=1080, height=1920,
#                 accept_group_id=False,   # already resolved to a look id
#             )
#         else:
#             video_id = avatar_heygen.create_talking_photo_video_from_text(
#                 talking_photo_id=look_id,
#                 input_text=sr.final_script,
#                 voice_id=picked_voice,
#                 title=f"{sr.icon_or_topic} · req#{sr.id}",
#                 width=1080, height=1920,
#             )
#     except HTTPError as e:
#         # --- Fallback: ElevenLabs TTS -> upload audio -> render via audio asset ---
#         el_voice = getattr(sr.avatar, "elevenlabs_voice_id", None)
#         mp3_bytes = tts_elevenlabs.synthesize_tts_bytes(sr.final_script, el_voice)  # see adapter change below
#         audio_asset_id = avatar_heygen.upload_audio_asset(mp3_bytes)

#         if character_type == "avatar":
#             video_id = avatar_heygen.create_avatar_video_from_audio(
#                 avatar_id=look_id,
#                 audio_asset_id=audio_asset_id,
#                 title=f"{sr.icon_or_topic} · req#{sr.id}",
#                 width=1080, height=1920,
#                 accept_group_id=False,
#             )
#         else:
#             video_id = avatar_heygen.create_talking_photo_video_from_audio(
#                 talking_photo_id=look_id,
#                 audio_asset_id=audio_asset_id,
#                 title=f"{sr.icon_or_topic} · req#{sr.id}",
#                 width=1080, height=1920,
#             )

#     # Wait and persist
#     st = avatar_heygen.wait_for_video(video_id, timeout_sec=900, poll_sec=10)
#     if st.get("status") != "completed":
#         sr.status = "Assembling"
#         sr.qc_json = {**(sr.qc_json or {}), "heygen_status": st}
#         sr.save()
#         return st

#     sr.asset_url = st.get("video_url", "")
#     sr.edit_url = avatar_heygen.get_share_url(video_id) or ""
#     sr.status = "Rendered"
#     sr.updated_at = timezone.now()
#     sr.save()
#     return {"video_url": sr.asset_url, "share_url": sr.edit_url}


@shared_task
def task_kickoff_chain(sr_id: int):
    flow = chain(
        # task_generate_script.si(sr_id),
        task_render_avatar.si(sr_id),
        task_assemble_template.si(sr_id),
        task_push_drive.si(sr_id),
        task_generate_captions.si(sr_id),
        task_sync_airtable.si(sr_id),
        task_schedule.si(sr_id),
    )
    flow.delay()
    return {"pipeline": "queued", "sr_id": sr_id}

# ====================== Batch Paragraph Generation ======================



def _abs_and_ensure_parent(relpath: str) -> str:
    """
    Returns absolute path for a storage-relative file and ensures its parent dir exists.
    Only for FileSystemStorage. (If you switch to S3, switch to BytesIO+default_storage.save.)
    """
    abs_path = default_storage.path(relpath)
    parent = os.path.dirname(abs_path)
    os.makedirs(parent, exist_ok=True)
    return abs_path

@shared_task(bind=True, autoretry_for=(Exception,), retry_backoff=True, retry_kwargs={"max_retries": 3})
def process_row_task(self, row_dict: dict) -> dict:
    """
    row_dict: {"row": int, "icon": str, "category": str, "notes": str}
    Returns:  {"row", "icon", "category", "notes", "paragraph", "ssml"}
    """
    icon = (row_dict.get("icon") or "").strip()
    category = (row_dict.get("category") or "").strip()
    notes = (row_dict.get("notes") or "").strip()

    prompt = build_prompt(icon=icon, notes=notes, category=category)
    raw = call_openai_for_paragraph_and_ssml(prompt)
    paragraph, ssml = parse_openai_json(raw)

    return {
        "row": row_dict["row"],
        "icon": icon,
        "category": category,
        "notes": notes,
        "paragraph": paragraph,
        "ssml": ssml,
    }


@shared_task(bind=True)
def save_batch_task(
    self,
    batch_results: list,
    job_id: str,
    batch_no: int,
    results_path: str,
    # NEW (optional) — for Google Sheet write-back
    mode: Optional[str] = None,
    sheet_id: Optional[str] = None,
    sheet_name: Optional[str] = None,
) -> dict:
    """
    Appends batch_results to an XLSX file (always),
    and if mode == "google_sheet" and sheet_id/sheet_name are provided,
    writes Paragraph + SSML back to the Google Sheet for the exact rows.
    """
    # Sort for deterministic append order
    batch_results = sorted(batch_results, key=lambda x: x.get("row", 0))

    # ---- 1) Append to results workbook
    abs_results = _abs_and_ensure_parent(results_path)
    if not os.path.exists(abs_results):
        wb = Workbook()
        ws = wb.active
        ws.append(["row", "icon", "category", "notes", "paragraph", "ssml"])
    else:
        wb = load_workbook(abs_results)
        ws = wb.active

    for item in batch_results:
        ws.append([
            item.get("row", ""),
            item.get("icon", ""),
            item.get("category", ""),
            item.get("notes", ""),
            item.get("paragraph", ""),
            item.get("ssml", ""),
        ])
    wb.save(abs_results)

    # ---- 2) Optional Google Sheet write-back
    if (mode or "").lower() == "google_sheet" and sheet_id and sheet_name and batch_results:
        try:
            sheets = _sheets_client()
            # Ensure headers and get column map
            colmap = _ensure_sheet_headers_and_map(sheets, sheet_id, sheet_name, ["Paragraph", "SSML"])
            p_col = colmap["Paragraph"]
            s_col = colmap["SSML"]

            # Stage ranges to batchUpdate
            data = []
            for item in batch_results:
                rownum = int(item.get("row", 0) or 0)
                if rownum < 2:
                    continue
                paragraph = (item.get("paragraph") or "").strip()
                ssml = (item.get("ssml") or "").strip()
                p_a1 = f"{_a1_col(p_col)}{rownum}"
                s_a1 = f"{_a1_col(s_col)}{rownum}"
                data.append({"range": f"{sheet_name}!{p_a1}", "values": [[paragraph]]})
                data.append({"range": f"{sheet_name}!{s_a1}", "values": [[ssml]]})

            if data:
                sheets.values().batchUpdate(
                    spreadsheetId=sheet_id,
                    body={"valueInputOption": "RAW", "data": data},
                ).execute()
        except Exception as e:
            logger.exception(f"Google Sheet write-back failed: {e}")

    logger.info(f"[Batch {batch_no}] Saved {len(batch_results)} rows to {results_path}")
    return {"saved_batch": batch_no, "count": len(batch_results)}

@shared_task(bind=True)
def orchestrate_paragraphs_job(
    self,
    # local file mode (backward compatible)
    file_path: Optional[str] = None,
    sheet: Optional[str] = None,
    batch_size: int = 25,
    # google sheet mode
    mode: str = "local_file",                 # "local_file" | "google_sheet"
    sheet_public_url: Optional[str] = None,   # required if mode == "google_sheet"
    sheet_id: Optional[str] = None,           # optional: for write-back
    sheet_name: Optional[str] = None,         # optional: default "Sheet1"
) -> dict:
    """
    Orchestrates paragraph generation in batches from either:
      - local .xlsx (mode='local_file'), or
      - Google Sheet public CSV (mode='google_sheet').

    Returns: { job_id, results: <xlsx path>, batches: <int>, mode: <mode> }
    """
    job_id = self.request.id

    # Prepare results workbook with headers
    results_rel = f"{RESULTS_DIR}/{job_id}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["row", "icon", "category", "notes", "paragraph", "ssml"])
    abs_results = _abs_and_ensure_parent(results_rel)
    wb.save(abs_results)

    # --- Row sources ---
    def google_sheet_rows() -> Iterable[Dict]:
        if not sheet_public_url:
            raise ValueError("sheet_public_url is required for mode='google_sheet'")
        resp = requests.get(sheet_public_url, timeout=30)
        resp.raise_for_status()
        df = pd.read_csv(io.StringIO(resp.text))

        cols_lower = {c.lower().strip(): c for c in df.columns}
        def pick(*cands):
            for c in cands:
                if c in cols_lower:
                    return cols_lower[c]
            return None

        col_icon = pick("icon name", "icon", "name")
        col_cat = pick("category",)
        col_notes = pick("notes", "note", "description")
        if not (col_icon or col_cat or col_notes):
            raise ValueError("CSV is missing required columns: (Icon Name | Category | Notes)")

        for idx, r in df.iterrows():
            sheet_rownum = idx + 2  # row 1 = header
            icon = str(r.get(col_icon, "")).strip() if col_icon else ""
            cat = str(r.get(col_cat, "")).strip() if col_cat else ""
            notes = str(r.get(col_notes, "")).strip() if col_notes else ""
            if not (icon or cat or notes):
                continue
            yield {"row": sheet_rownum, "icon": icon, "category": cat, "notes": notes}

    def local_file_rows() -> Iterable[Dict]:
        if not file_path:
            raise ValueError("file_path is required for mode='local_file'")
        abs_path = default_storage.path(file_path)
        for r in iter_rows_streaming(abs_path, sheet=sheet):
            yield r

    # choose source
    if (mode or "").lower() == "google_sheet":
        row_iter = google_sheet_rows()
        mode = "google_sheet"
    else:
        row_iter = local_file_rows()
        mode = "local_file"

    # --- batching ---
    def _chunker(it: Iterable[Dict], n: int) -> Iterable[List[Dict]]:
        buf: List[Dict] = []
        for item in it:
            buf.append(item)
            if len(buf) == n:
                yield buf
                buf = []
        if buf:
            yield buf

    total_batches = 0
    for total_batches, rows in enumerate(_chunker(row_iter, batch_size), start=1):
        g = group(process_row_task.s(row) for row in rows)
        # Pass mode/sheet params to the callback so it can write back to Google Sheet
        chord(g)(save_batch_task.s(job_id, total_batches, results_rel, mode, sheet_id, sheet_name or "Sheet1"))

        logger.info(f"[{mode}] Scheduled batch {total_batches} for job {job_id}")
        self.update_state(state="PROGRESS", meta={"scheduled_batches": total_batches, "mode": mode})

    download_url = None
    try:
        download_url = default_storage.url(results_rel)  # e.g. /media/uploads/results/<job>.xlsx
    except Exception:
        pass

    return {
        "job_id": job_id,
        "results": results_rel,
        "download_url": download_url,   # <-- return this
        "batches": total_batches,
        "mode": mode,
    }
